from collections import defaultdict, Counter
from os import getcwd, system, listdir, makedirs
from os.path import isfile, join, dirname, abspath
from warnings import filterwarnings
from time import perf_counter

from Bio import SeqIO, BiopythonWarning
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from typing import NamedTuple

from src.classes import Binder, SeqLiab


"""
TO OD
    > Analysis report:
        - No of seqs
        - No of unique binders
        - Avr no of seq per binder
        - ...
    > Move analysis funcs to Binder methods
    ...
"""


class UniqueBinder(NamedTuple):
    rfs: str
    lib: str
    num: int
    recs: list


def print_frame(func):
    def wrapper(*args, **kwargs):
        system('cls')
        print('/ PureMantis \\'.center(80, '_'))
        func(*args, **kwargs)
        print(''.center(80, '‾'))
    return wrapper


def measure_time(func):
    def wrapper(*args, **kwargs):
        start_time = perf_counter()
        result = func(*args, **kwargs)
        exec_time = perf_counter() - start_time
        print(f'Analyzed {len(result)} records in {exec_time:.3f} s\n')
        return result
    return wrapper


@print_frame
def main():

    filterwarnings('ignore', category=BiopythonWarning, module='Bio')

    mypath = getcwd()
    formats = ['fasta', 'fastq']
    files = get_files(mypath, formats)

    # Load Excel template
    wb = load_template()

    # Analyze fasta files
    results = parse_files(files)
    unique_binders = get_unique_binders(results)

    # Write results to template
    write_results(wb, results)
    write_ubs(wb, unique_binders)

    # Create folder
    file_name = input("Save as: ")
    file_dir = join(mypath, file_name)
    file_path = join(file_dir, file_name + ".xlsx")
    makedirs(f'{file_dir}', exist_ok=True)

    # Save results
    save_workbook(wb, file_path)

    # Sort scf files
    sort_scf_files(unique_binders, mypath, file_dir)

    # Move fasta
    move_files(files, file_dir)


def get_files(mypath: str, format: list, subdir: str = None, abspath: bool = True) -> list:
    if subdir is not None:
        mypath = join(mypath, subdir)
    try:
        files = [file_path if abspath else f for f in listdir(mypath)
                if isfile(file_path := join(mypath, f))
                and f.rpartition('.')[-1].lower() in format]
        return files
    except FileNotFoundError:
        print(f'---< {subdir} folder not found >---'.center(80))
        return []


def sort_scf_files(unique_binders: list, path: str, file_dir: str) -> None:
    scf_files = get_files(path, 'scf', 'scf_files', False)
    if scf_files:
        for i, binder in enumerate(unique_binders, 1):
            makedirs(f'{file_dir}\\Variant {i}', exist_ok=True)
            for seq_id, seq_quality in binder.recs:
                if '_(reversed)' in seq_id:
                    seq_id = seq_id.replace('_(reversed)', '')
                if '.SCF' not in seq_id:
                    seq_id = seq_id + '.SCF'
                if seq_id in scf_files:
                    system(f'move ".\\scf_files\\{seq_id}" "{file_dir}\\Variant {i}\\" >nul')
                    scf_files.remove(seq_id)
        for file in scf_files:
            makedirs(f'{file_dir}\\Not classified', exist_ok=True)
            system(f'move ".\\scf_files\\{file}" "{file_dir}\\Not classified\\" >nul')
    else:
        print("---< No scf files detected >---".center(80))


def load_template():
    template_dir = dirname(abspath(__file__))
    template_name = 'rf_template.xlsx'
    template_path = join(template_dir, 'templates', template_name)

    return load_workbook(template_path)


@measure_time
def parse_files(files) -> list:
    results = []
    for file in files:
        ext = file.rpartition('.')[-1]
        with open(file, 'r') as handle:
            for record in SeqIO.parse(handle, ext):
                binder = Binder(record)
                result = binder.analyze_binder()
                results.append(result)
    # Sort: RFs: ascending | quality (HQ): descending
    results.sort(key=lambda x: (
        list(map(str, x.rfs)),
        (-x.quality[0] if x.quality else 0))
        )
    return results


def get_unique_binders(results: list) -> list:

    unique_binders = defaultdict(list)

    for result in results:
        if '-' not in result.rfs:
            ub_key = ('_'.join(map(str, result.rfs)), result.lib)
            unique_binders[ub_key] += [(result.seq_id, result.quality)]
    return sort_unique_binders(unique_binders)


def get_pI(rfs) -> float:

    # Calculates charge of aa at given pH based on pKa
    def charge(ph, pka):
        return 1 / (10 ** (ph-pka)+1)

    if isinstance(rfs, list):
        rfs = ''.join(map(str, rfs)).upper()

    aa_count = Counter(rfs)

    # Set range of pI search and start pH
    pH_min, pH, pH_max = 2, 8, 14.0

    base_pka = {'K': 10.00,
                'R': 12.00,
                'H':  5.98}
    acid_pka = {'D':  4.05,
                'E':  4.45,
                'C':  9.00,
                'Y': 10.00}
    n_term_dict = {'A': 7.59,
                   'M': 7.00,
                   'S': 6.93,
                   'P': 8.36,
                   'T': 6.82,
                   'V': 7.44,
                   'E': 7.70}
    c_term_dict = {'D': 4.55,
                   'E': 4.75}

    n_pka = n_term_dict.get(rfs[0], 7.7)
    c_pka = c_term_dict.get(rfs[-1], 3.55)

    p_charge = sum([charge(pH, base_pka[aa]) * aa_count[aa] for aa in base_pka.keys()]) + charge(pH, n_pka)
    n_charge = sum([charge(acid_pka[aa], pH) * aa_count[aa] for aa in acid_pka.keys()]) + charge(c_pka, pH)
    net_charge = p_charge - n_charge

    while abs(net_charge) > 0.001:
        if net_charge > 0:
            pH_min = pH
        else:
            pH_max = pH

        pH = (pH_min + pH_max) / 2

        p_charge = sum([charge(pH, base_pka[aa]) * aa_count[aa] for aa in base_pka.keys()]) + charge(pH, n_pka)
        n_charge = sum([charge(acid_pka[aa], pH) * aa_count[aa] for aa in acid_pka.keys()]) + charge(c_pka, pH)
        net_charge = p_charge - n_charge

    return round(pH, 2)


def write_results(wb: Workbook, results: list) -> None:
    # Sort based on RFs > library > mutations > STOP mutations > AA sequence
    ws_rfs = wb["RF analysis"]

    # TODO refactor
    start_row = 4
    for i, result in enumerate(results, start_row):
        # Add sequence ID
        ws_rfs.cell(column=2, row=i).value = result.seq_id
        # Add RFs
        for j in range(6):
            ws_rfs.cell(column=3+j, row=i).value = str(result.rfs[j])
        # Add detected library
        ws_rfs.cell(column=10, row=i).value = result.lib
        # Add HQ%, min phred score and it's position for fastq records
        if result.quality:
            for j in range(3):
                ws_rfs.cell(column=11+j, row=i).value = result.quality[j]
        # Add mutations of NRFs
        for j in range(8):
            ws_rfs.cell(column=14+j, row=i).value = result.frame_muts[j]
        # Add stop codon detection
        for j in range(3):
            ws_rfs.cell(column=22+j, row=i).value = result.stop_muts[j]
        # Add full AA sequence
        ws_rfs.cell(column=25, row=i).value = str(result.seq_aa)

def sort_unique_binders(unique_binders: defaultdict) -> list:
    return sorted(
        [UniqueBinder(ub_key,  # rfs
                      ub_lib,  # lib
                      len(seq_records),  # num
                      seq_records)  # recs: [(seq_id, seq_quality), ...]
        for (ub_key, ub_lib), seq_records in unique_binders.items()],
        key=lambda x: x.num,
        reverse=True)


def write_ubs(wb: Workbook, unique_binders: list) -> None:
    # Change dict into list and add number of IDs per RF
    # ub: ((rfs_merged, lib), [(res_id, res_quality), ...])

    ws_ubs = wb["Unique binders"]

    sliab_cols = {
            'Glycosylation': 13,
            'Isomerization': 14,
            'Fragmentation': 15,
            'Extra Cysteine': 16,
            'Deamidation CDR (High)': 17,
            'Hydrolysis': 18,
            'Cleavage': 19,
            'Deamidation CDR (Med)': 20,
            'Integrin binding aVb3': 21,
            'Integrin binding a4b1': 22,
            'Integrin binding a2b1': 23,
            'CD11c/CD18 binding': 24,
            'Hydrophobicity': 25,
            'Oxidation': 26,
            'Deamidation CDR (Low)': 27,
        }

    rf_dict = {
                1: 'RFH1',
                2: 'RFH2',
                3: 'RFH3',
                4: 'RFL1',
                5: 'RFL2',
                6: 'RFL3',
               }

    # Sort IDs by quality
    for binder in unique_binders:
        binder.recs.sort(key=lambda x: x[1], reverse=True)

    # Write RFs of given variant and IDs sharing it
    start_row = 5
    for i, ub in enumerate(unique_binders, start_row):

        # Write library
        ws_ubs.cell(column=2, row=i).value = ub.lib

        # Write randomized fragments
        rfs_split = ub.rfs.split("_")
        for j, rf in enumerate(rfs_split):
            ws_ubs.cell(column=3+j, row=i).value = rf

        # Write umber of seqs
        ws_ubs.cell(column=9, row=i).value = ub.num

        # Write pH
        ws_ubs.cell(column=11, row=i).value = get_pI(ub.rfs)

        # Analyze sequence liabilities
        sliab = SeqLiab(ub.rfs)

        # Write liability score
        ws_ubs.cell(column=12, row=i).value = sliab.get_score()

        # Write sequence liabilities
        for liab, liab_count, liab_list in sliab.get_summary():
            text = '\n'.join(map(lambda x: f'{rf_dict.get(x[0])}: {x[1]}', liab_list))
            comment = Comment(text=text,
                              author='',
                              height=17.5 * len(liab_list),
                              width=85)
            ws_ubs.cell(column=sliab_cols[liab], row=i).value = liab_count
            ws_ubs.cell(column=sliab_cols[liab], row=i).comment = comment

        # Write record names
        for j, rec in enumerate(ub.recs):  # Seq IDs
            ws_ubs.cell(column=28+j, row=i).value = rec[0]


def save_workbook(wb: Workbook, file_path: str) -> None:
    while True:
        try:
            wb.save(filename=file_path)
            break
        except PermissionError:
            print('---< Unable to save: file in use - close it and try again >---'.center(80))
            system('pause')


def move_files(files: list, target_dir: str) -> None:
    for file in files:
        system(f'move "{file}" "{target_dir}" >nul')


if __name__ == "__main__":
    main()
    
    # seq = "ATGAAATACCTATTGCCTACGGCAGCCGCTGGATTGTTATTACTCGCGGCCCAGCCGGCCATGGCCGAGGTGCAGCTGTTGGAGTCTGGGGGAGGCTTGGTACAGCCTGGGGGGTCCCTGAGACTCTCCTGTGCAGCCTCTTAGTTCACCTTTAGCAGCTATGCCATGAGCTGGGTCCGCCAGGCTCCAGGGAAGGGGCTGGAGTGGGTGTCAGCTATTAGTGGTAGTGGTGGTAGCACATACTACGCAGACTCCGTGAAGGGCCGGTTCACCATCTCCAGAGACAATTCCAAGAACACGCTGTATCTGCAAATGAACAGCCTGAGAGCCGAGGACACGGCCGTATATTACTGTGCGAAAGAAGTCGGATATTGTAGTAGTACCAGCTTTGACCCCTGGGGCCAGGGAACCCTGGTCACCGTGTCCTCAGGTGGAGGCGGTTCAGGCGGAGGTGGCAGCGGCGGTGGCGGGTCGACGGAAATTGTGTTGACGCAGTCTCCAGGCACCCTGTCTTTGTCTCCAGGGGAAAGAGCCACCCTCTCCTGCAGGGCCAGTCAGAGTGTTAGCAGCAGCTACTTAGCCTGGTACCAGCAGAAACCTGGCCAGGCTCCCAGGCTCCTCATCTATGGTGCATCCAGCAGGGCCACTGGCATCCCAGACAGGTTCAGTGGCAGTGGGTCTGGGACAGACTTCACTCTCACCATCAGCAGACTGGAGCCTGAAGATTTTGCAGTGTATTACTGTCAGCAGTATGGTAGCTCACGGGGGTACTTTGGCCAGGGGACCAAGCTGGAGATCAAACGAGCGGCCGCAGGGGCCGCAGAACAAAAACTCATCTCAGAAGAGGATCTGGGAGACGCG"
    # binder = Binder.seq(seq)
    # res = measure_time(binder.analyze_binder)()
    # print(binder.rfs)
    
